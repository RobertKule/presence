# core/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
from django.db.models import Count, Q
from django.utils import timezone
from django.core.paginator import Paginator

import openpyxl
from openpyxl.styles import Font
from io import BytesIO
import json
import openpyxl
from xhtml2pdf import pisa
from django.core.serializers.json import DjangoJSONEncoder
import csv
import io

# Import des formulaires
from .forms import (
    SignUpForm, CoursForm, SeanceForm,
    EnseignantForm, ClasseForm, EtudiantForm, ImportEtudiantsForm
)

# Import des modèles
from .models import (
    Cours, Seance, Classe, User,
    Presence, Etudiant
)


# -------------------
# UTILS & DECORATEURS
# -------------------

def admin_required(user):
    """Vérifie si l'utilisateur est administrateur"""
    return user.is_authenticated and user.role == "admin"

def enseignant_required(user):
    """Vérifie si l'utilisateur est enseignant"""
    return user.is_authenticated and user.role == "enseignant"

# -------------------
# AUTHENTIFICATION
# -------------------

def signup_view(request):
    """Vue d'inscription"""
    if request.user.is_authenticated:
        return redirect('/')
        
    if request.method == "POST":
        form = SignUpForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, "Inscription réussie !")
            return redirect('/')
    else:
        form = SignUpForm()
    return render(request, "core/signup.html", {"form": form})

def login_view(request):
    """Vue de connexion"""
    if request.user.is_authenticated:
        return redirect('/')
        
    if request.method == "POST":
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            messages.success(request, f"Bienvenue {user.username} !")
            return redirect('/')
        else:
            messages.error(request, "Identifiants incorrects")
    else:
        form = AuthenticationForm()
    return render(request, "core/login.html", {"form": form})

def logout_view(request):
    """Vue de déconnexion"""
    logout(request)
    messages.info(request, "Vous avez été déconnecté")
    return redirect('core:login')

# -------------------
# DASHBOARDS
# -------------------
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from .models import Cours, Seance, Etudiant, Presence, User

def home(request):
    """Page d'accueil publique avec statistiques réelles"""
    if request.user.is_authenticated:
        if request.user.role == "admin":
            return redirect('core:admin_dashboard')
        return redirect('core:dashboard')

    # Statistiques réelles
    stats = {
        'total_appels': Presence.objects.count(),
        'total_etudiants': Etudiant.objects.count(),
        'total_cours': Cours.objects.count(),
        'total_heures_gagnees': (Presence.objects.count() // 60)  # 1 minute par appel manuel → heures gagnées
    }

    return render(request, 'core/home.html', {'stats': stats})
@login_required
def dashboard(request):
    """Dashboard principal selon le rôle"""
    if request.user.role == "admin":
        return redirect('core:admin_dashboard')
    
    # Dashboard enseignant
    cours = request.user.cours_enseignant.all().order_by('-created_at')[:5]
    seances_recentes = Seance.objects.filter(
        cours__enseignant=request.user
    ).order_by('-date', '-heure_debut')[:5]
    
    # Statistiques rapides
    stats = {
        'total_cours': request.user.cours_enseignant.count(),
        'total_seances': Seance.objects.filter(cours__enseignant=request.user).count(),
        'seances_aujourdhui': Seance.objects.filter(
            cours__enseignant=request.user,
            date=timezone.now().date()
        ).count()
    }
    
    return render(request, "core/dashboard.html", {
        "cours": cours,
        "seances_recentes": seances_recentes,
        "stats": stats
    })

@login_required
@user_passes_test(admin_required)
def admin_dashboard(request):
    """Dashboard administrateur"""
    enseignants = User.objects.filter(role="enseignant")
    classes = Classe.objects.all()
    etudiants = Etudiant.objects.all()
    cours = Cours.objects.all()
    
    stats = {
        'total_enseignants': enseignants.count(),
        'total_classes': classes.count(),
        'total_etudiants': etudiants.count(),
        'total_cours': cours.count(),
        'presences_aujourdhui': Presence.objects.filter(
            seance__date=timezone.now().date()
        ).count()
    }
    
    return render(request, "core/admin_dashboard.html", {
        "enseignants": enseignants,
        "classes": classes,
        "etudiants": etudiants,
        "cours": cours,
        "stats": stats
    })

# -------------------
# GESTION DES COURS
# -------------------

@login_required
@user_passes_test(enseignant_required)
def mes_cours(request):
    """Liste des cours de l'enseignant"""
    cours_list = request.user.cours_enseignant.all().order_by('-created_at')
    
    # Pagination
    paginator = Paginator(cours_list, 10)
    page_number = request.GET.get('page')
    cours = paginator.get_page(page_number)
    
    return render(request, "core/mes_cours.html", {"cours": cours})

@login_required
@user_passes_test(enseignant_required)
def cours_detail(request, pk):
    """Détail d'un cours avec statistiques"""
    cours = get_object_or_404(Cours, pk=pk, enseignant=request.user)
    seances = cours.seances.all().order_by('-date')
    etudiants = cours.classe.etudiants.all()
    
    # Statistiques détaillées
    stats = {
        'total_seances': seances.count(),
        'total_etudiants': etudiants.count(),
        'presences_par_etudiant': [],
        'stats_globales': {
            'present': Presence.objects.filter(seance__cours=cours, statut='present').count(),
            'absent': Presence.objects.filter(seance__cours=cours, statut='absent').count(),
            'retard': Presence.objects.filter(seance__cours=cours, statut='retard').count(),
            'motif': Presence.objects.filter(seance__cours=cours, statut='motif').count(),
        }
    }
    
    for etu in etudiants:
        presences = etu.presences.filter(seance__cours=cours)
        total_presences = presences.count()
        
        stats['presences_par_etudiant'].append({
            'etudiant': etu,
            'present': presences.filter(statut='present').count(),
            'retard': presences.filter(statut='retard').count(),
            'absent': presences.filter(statut='absent').count(),
            'motif': presences.filter(statut='motif').count(),
            'taux_presence': (presences.filter(statut__in=['present', 'retard']).count() / total_presences * 100) if total_presences > 0 else 0
        })
    
    return render(request, "core/cours_detail.html", {
        "cours": cours,
        "seances": seances,
        "stats": stats
    })

@login_required
@user_passes_test(enseignant_required)
def cours_create(request):
    """Création de cours par l'enseignant"""
    if request.method == "POST":
        form = CoursForm(request.POST, user=request.user)
        if form.is_valid():
            cours = form.save(commit=False)
            cours.enseignant = request.user
            cours.save()
            messages.success(request, f"Cours '{cours.nom}' créé avec succès !")
            return redirect("core:mes_cours")
    else:
        form = CoursForm(user=request.user)
    
    return render(request, "core/cours_form.html", {
        "form": form,
        "title": "Créer un nouveau cours",
        "action_url": "cours_create"
    })

@login_required
@user_passes_test(enseignant_required)
def cours_update(request, pk):
    """Modification de cours par l'enseignant"""
    cours = get_object_or_404(Cours, pk=pk, enseignant=request.user)
    
    if request.method == "POST":
        form = CoursForm(request.POST, instance=cours, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, f"Cours '{cours.nom}' modifié avec succès !")
            return redirect("core:mes_cours")
    else:
        form = CoursForm(instance=cours, user=request.user)
    
    return render(request, "core/cours_form.html", {
        "form": form,
        "title": f"Modifier le cours : {cours.nom}",
        "action_url": "cours_update"
    })

@login_required
@user_passes_test(enseignant_required)
def cours_delete(request, pk):
    """Suppression de cours avec confirmation"""
    cours = get_object_or_404(Cours, pk=pk, enseignant=request.user)
    
    if request.method == "POST":
        nom_cours = cours.nom
        cours.delete()
        messages.success(request, f"Cours '{nom_cours}' supprimé avec succès !")
        return redirect("core:mes_cours")
    
    return render(request, "core/cours_confirm_delete.html", {"cours": cours})

# -------------------
# GESTION DES SÉANCES
# -------------------

@login_required
@user_passes_test(enseignant_required)
def seance_list(request):
    """Liste des séances de l'enseignant"""
    seances_list = Seance.objects.filter(cours__enseignant=request.user).order_by('-date', '-heure_debut')
    
    # Filtrage par cours si spécifié
    cours_id = request.GET.get('cours')
    if cours_id:
        seances_list = seances_list.filter(cours_id=cours_id)
    
    # Pagination
    paginator = Paginator(seances_list, 10)
    page_number = request.GET.get('page')
    seances = paginator.get_page(page_number)
    
    cours_options = request.user.cours_enseignant.all()
    
    return render(request, "core/seance_list.html", {
        "seances": seances,
        "cours_options": cours_options
    })

@login_required
@user_passes_test(enseignant_required)
def seance_create(request):
    """Création d'une séance"""
    if request.method == "POST":
        form = SeanceForm(request.POST, user=request.user)
        if form.is_valid():
            seance = form.save()
            messages.success(request, f"Séance du {seance.date} créée avec succès !")
            return redirect("core:seance_list")
    else:
        form = SeanceForm(user=request.user)
    
    return render(request, "core/seance_form.html", {
        "form": form,
        "title": "Créer une nouvelle séance"
    })

@login_required
@user_passes_test(enseignant_required)
def seance_update(request, pk):
    """Modification d'une séance"""
    seance = get_object_or_404(Seance, pk=pk, cours__enseignant=request.user)
    
    if request.method == "POST":
        form = SeanceForm(request.POST, instance=seance, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, f"Séance du {seance.date} modifiée avec succès !")
            return redirect("core:seance_list")
    else:
        form = SeanceForm(instance=seance, user=request.user)
    
    return render(request, "core/seance_form.html", {
        "form": form,
        "title": f"Modifier la séance du {seance.date}"
    })

@login_required
@user_passes_test(enseignant_required)
def seance_delete(request, pk):
    """Suppression d'une séance"""
    seance = get_object_or_404(Seance, pk=pk, cours__enseignant=request.user)
    
    if request.method == "POST":
        date_seance = seance.date
        seance.delete()
        messages.success(request, f"Séance du {date_seance} supprimée avec succès !")
        return redirect("core:seance_list")
    
    return render(request, "core/seance_confirm_delete.html", {"seance": seance})

# -------------------
# GESTION DES PRÉSENCES
# -------------------
@login_required
@user_passes_test(enseignant_required)
def appel_presence(request, seance_id):
    """Appel de présence pour une séance"""
    seance = get_object_or_404(Seance, pk=seance_id, cours__enseignant=request.user)
    etudiants = Etudiant.objects.filter(classe=seance.cours.classe).order_by('nom', 'prenom')
    
    # Récupérer les présences existantes
    presences_existantes = Presence.objects.filter(seance=seance)
    presences_dict = {str(p.etudiant_id): p.statut for p in presences_existantes}
    
    # Calculer le taux de présence
    total_presences = presences_existantes.count()
    if total_presences > 0:
        presentes = presences_existantes.filter(statut__in=['present', 'retard']).count()
        taux_presence = (presentes / total_presences) * 100
    else:
        taux_presence = 0
    
    # Définir les statuts avec leurs icônes
    statuts = [
        ('present', 'fa-check'),
        ('retard', 'fa-clock'),
        ('absent', 'fa-times'),
        ('motif', 'fa-info-circle')
    ]
    
    if request.method == "POST":
        # Traitement de l'appel
        for etu in etudiants:
            statut = request.POST.get(f"statut_{etu.id}", "absent")
            Presence.objects.update_or_create(
                etudiant=etu,
                seance=seance,
                defaults={"statut": statut}
            )
        
        messages.success(request, f"Appel de présence enregistré pour la séance du {seance.date} !")
        return redirect("core:seance_list")
    
    return render(request, "core/appel_presence.html", {
        "seance": seance,
        "etudiants": etudiants,
        "presences": json.dumps(presences_dict),
        "statuts": statuts,
        "taux_presence": taux_presence
    })
@login_required
@user_passes_test(enseignant_required)
def ajouter_etudiant_rapide(request, seance_id):
    """Ajouter rapidement un étudiant depuis l'appel de présence"""
    seance = get_object_or_404(Seance, pk=seance_id, cours__enseignant=request.user)
    
    if request.method == "POST":
        matricule = request.POST.get('matricule')
        nom = request.POST.get('nom')
        prenom = request.POST.get('prenom')
        
        if not all([matricule, nom]):
            return JsonResponse({'success': False, 'error': 'Matricule et nom requis'})
        
        try:
            # Vérifier si l'étudiant existe déjà
            etudiant, created = Etudiant.objects.get_or_create(
                matricule=matricule,
                defaults={
                    'nom': nom,
                    'prenom': prenom or '',
                    'classe': seance.cours.classe
                }
            )
            
            if not created:
                # Si l'étudiant existe mais dans une autre classe, on le déplace
                if etudiant.classe != seance.cours.classe:
                    etudiant.classe = seance.cours.classe
                    etudiant.save()
            
            return JsonResponse({
                'success': True,
                'etudiant': {
                    'id': etudiant.id,
                    'nom': etudiant.nom,
                    'prenom': etudiant.prenom,
                    'matricule': etudiant.matricule
                },
                'created': created
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})

from django.http import JsonResponse

from django.db import transaction

def presence_rapide(request):
    if request.method == "POST":
        etudiant_id = request.POST.get("etudiant_id")
        seance_id = request.POST.get("seance_id")
        statut = request.POST.get("statut")

        try:
            with transaction.atomic():  # tout est atomique
                presence, created = Presence.objects.update_or_create(
                    etudiant_id=etudiant_id,
                    seance_id=seance_id,
                    defaults={"statut": statut}
                )
            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

# -------------------
# STATISTIQUES & RAPPORTS
# -------------------

@login_required
@user_passes_test(enseignant_required)
def statistiques(request):
    """Statistiques des présences"""
    cours_list = request.user.cours_enseignant.all()
    cours_id = request.GET.get('cours')
    
    if cours_id:
        cours_selected = get_object_or_404(Cours, id=cours_id, enseignant=request.user)
        cours_list = [cours_selected]
    else:
        cours_selected = None
    
    stats_globales = []
    
    for cours in cours_list:
        total_seances = cours.seances.count()
        etudiants = cours.classe.etudiants.all()
        
        stats_cours = {
            "cours": cours,
            "total_seances": total_seances,
            "global": {
                "present": 0,
                "absent": 0,
                "retard": 0,
                "motif": 0,
            },
            "etudiants_stats": []
        }
        
        for etu in etudiants:
            present = etu.presences.filter(seance__cours=cours, statut="present").count()
            absent = etu.presences.filter(seance__cours=cours, statut="absent").count()
            retard = etu.presences.filter(seance__cours=cours, statut="retard").count()
            motif = etu.presences.filter(seance__cours=cours, statut="motif").count()
            
            stats_cours["global"]["present"] += present
            stats_cours["global"]["absent"] += absent
            stats_cours["global"]["retard"] += retard
            stats_cours["global"]["motif"] += motif
            
            stats_cours["etudiants_stats"].append({
                "etudiant": etu,
                "present": present,
                "absent": absent,
                "retard": retard,
                "motif": motif,
                "taux_presence": (present + retard) / total_seances * 100 if total_seances > 0 else 0
            })
        
        stats_globales.append(stats_cours)
    
    return render(request, "core/statistiques.html", {
        "stats": stats_globales,
        "cours_selected": cours_selected,
        "cours_options": request.user.cours_enseignant.all(),
        "stats_json": json.dumps([{
            'cours': s['cours'].nom,
            'classe': s['cours'].classe.nom,
            'total_seances': s['total_seances'],
            'global': s['global'],
            'etudiants_stats': [{
                'etudiant': f"{es['etudiant'].nom} {es['etudiant'].prenom or ''}",
                'present': es['present'],
                'absent': es['absent'],
                'retard': es['retard'],
                'motif': es['motif'],
                'taux_presence': es['taux_presence']
            } for es in s['etudiants_stats']]
        } for s in stats_globales], cls=DjangoJSONEncoder)
    })

@login_required
@user_passes_test(enseignant_required)
def export_excel(request, cours_id):
    """Export Excel des présences"""
    cours = get_object_or_404(Cours, id=cours_id, enseignant=request.user)
    etudiants = cours.classe.etudiants.all().order_by('nom', 'prenom')
    seances = cours.seances.all().order_by("date")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Présences {cours.nom}"
    
    # En-têtes
    headers = ["Matricule", "Nom", "Prénom"] + [s.date.strftime("%d/%m/%Y") for s in seances] + ["Total Présent", "Total Absent", "Taux Présence"]
    ws.append(headers)
    
    # Données des étudiants
    for etu in etudiants:
        row = [etu.matricule, etu.nom, etu.prenom or ""]
        total_present = 0
        
        for s in seances:
            presence = Presence.objects.filter(etudiant=etu, seance=s).first()
            statut = presence.statut if presence else "absent"
            row.append(statut)
            if statut in ["present", "retard"]:
                total_present += 1
        
        total_absences = seances.count() - total_present
        taux_presence = (total_present / seances.count() * 100) if seances.count() > 0 else 0
        
        row.extend([total_present, total_absences, f"{taux_presence:.1f}%"])
        ws.append(row)
    
    # Formatage
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Export HTTP
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="presences_{cours.nom}_{timezone.now().date()}.xlsx"'
    wb.save(response)
    return response

@login_required
@user_passes_test(enseignant_required)
def export_pdf_statistiques(request, cours_id):
    """Export PDF des statistiques"""
    cours = get_object_or_404(Cours, id=cours_id, enseignant=request.user)
    etudiants = cours.classe.etudiants.all().order_by('nom', 'prenom')
    seances = cours.seances.all().order_by("date")
    
    etudiants_stats = []
    total_present = total_absent = total_retard = total_motif = 0
    
    for etu in etudiants:
        present = etu.presences.filter(seance__cours=cours, statut="present").count()
        absent = etu.presences.filter(seance__cours=cours, statut="absent").count()
        retard = etu.presences.filter(seance__cours=cours, statut="retard").count()
        motif = etu.presences.filter(seance__cours=cours, statut="motif").count()
        
        total_present += present
        total_absent += absent
        total_retard += retard
        total_motif += motif
        
        taux_presence = (present + retard) / seances.count() * 100 if seances.count() > 0 else 0
        
        etudiants_stats.append({
            "etudiant": etu,
            "present": present,
            "absent": absent,
            "retard": retard,
            "motif": motif,
            "taux_presence": taux_presence
        })
    
    stats = {
        "cours": cours,
        "total_seances": seances.count(),
        "global": {
            "present": total_present,
            "absent": total_absent,
            "retard": total_retard,
            "motif": total_motif,
        },
        "etudiants_stats": etudiants_stats,
        "date_generation": timezone.now().date()
    }
    
    html_string = render_to_string("core/statistiques_pdf.html", {"stats": stats})
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="statistiques_{cours.nom}_{timezone.now().date()}.pdf"'
    
    pisa_status = pisa.CreatePDF(
        src=html_string, 
        dest=response,
        encoding='UTF-8'
    )
    
    if pisa_status.err:
        return HttpResponse("Erreur lors de la génération du PDF")
    
    return response

# -------------------
# GESTION DES ÉTUDIANTS
# -------------------
@login_required
@user_passes_test(enseignant_required)
def mes_etudiants(request):
    """Liste des étudiants des classes de l'enseignant"""
    classes_enseignant = Classe.objects.filter(cours__enseignant=request.user).distinct()
    classe_id = request.GET.get('classe')
    
    if classe_id:
        etudiants = Etudiant.objects.filter(classe_id=classe_id, classe__in=classes_enseignant)
        # Récupérer le nom de la classe sélectionnée
        classe_selected = int(classe_id)
        classe_nom = Classe.objects.get(id=classe_id).nom
    else:
        etudiants = Etudiant.objects.filter(classe__in=classes_enseignant)
        classe_selected = None
        classe_nom = "Toutes les classes"
    
    etudiants = etudiants.order_by('classe__nom', 'nom', 'prenom')
    
    # Pagination
    paginator = Paginator(etudiants, 20)
    page_number = request.GET.get('page')
    etudiants_page = paginator.get_page(page_number)
    
    return render(request, "core/mes_etudiants.html", {
        "etudiants": etudiants_page,
        "classes": classes_enseignant,
        "classe_selected": classe_selected,
        "classe_nom": classe_nom  
    })

@login_required
@user_passes_test(enseignant_required)
def importer_etudiants(request):
    """Import d'étudiants via fichier Excel/CSV"""
    classes_enseignant = Classe.objects.filter(cours__enseignant=request.user)
    
    if request.method == "POST":
        form = ImportEtudiantsForm(request.POST, request.FILES)
        form.fields['classe'].queryset = classes_enseignant
        
        if form.is_valid():
            fichier = request.FILES['fichier']
            classe = form.cleaned_data['classe']
            mode = form.cleaned_data.get('mode', 'create')
            
            try:
                if fichier.name.endswith('.xlsx'):
                    # Traitement Excel
                    wb = openpyxl.load_workbook(fichier)
                    ws = wb.active
                    lignes = list(ws.iter_rows(values_only=True))
                else:
                    # Traitement CSV
                    data = fichier.read().decode('utf-8')
                    io_string = io.StringIO(data)
                    lignes = list(csv.reader(io_string))
                
                etudiants_importes = 0
                etudiants_modifies = 0
                errors = []
                
                for i, row in enumerate(lignes[1:], start=2):  # Skip header
                    try:
                        if len(row) >= 2 and row[0] and row[1]:
                            matricule = str(row[0]).strip()
                            nom = str(row[1]).strip()
                            prenom = str(row[2]).strip() if len(row) > 2 else ""
                            
                            if mode == 'create':
                                etudiant, created = Etudiant.objects.get_or_create(
                                    matricule=matricule,
                                    defaults={'nom': nom, 'prenom': prenom, 'classe': classe}
                                )
                                if created:
                                    etudiants_importes += 1
                            else:
                                # Mode update
                                etudiant, created = Etudiant.objects.update_or_create(
                                    matricule=matricule,
                                    defaults={'nom': nom, 'prenom': prenom, 'classe': classe}
                                )
                                if created:
                                    etudiants_importes += 1
                                else:
                                    etudiants_modifies += 1
                    except Exception as e:
                        errors.append(f"Ligne {i}: {str(e)}")
                
                if errors:
                    messages.warning(request, f"Import terminé avec {len(errors)} erreurs")
                else:
                    messages.success(request, 
                        f"{etudiants_importes} étudiants importés, {etudiants_modifies} modifiés" 
                        if mode == 'update' else f"{etudiants_importes} étudiants importés"
                    )
                
                return redirect("core:mes_etudiants")
                
            except Exception as e:
                messages.error(request, f"Erreur lors de la lecture du fichier: {str(e)}")
    else:
        form = ImportEtudiantsForm()
        form.fields['classe'].queryset = classes_enseignant
    
    return render(request, "core/importer_etudiants.html", {"form": form})
from django.http import HttpResponse
from openpyxl import Workbook
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods

@login_required
@require_http_methods(["GET"])
def telecharger_modele_import(request):
    """Télécharger un modèle Excel pré-rempli pour l'import"""
    # Créer un classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Modèle Import"

    # En-têtes
    ws.append(["Matricule", "Nom", "Prénom"])

    # Exemple avec votre nom
    ws.append(["RK001", "KULE", "Robert"])

    # Style basique
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    # Ajuster la largeur des colonnes
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    # Sauvegarder en mémoire
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Créer la réponse HTTP
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="modele_import_etudiants.xlsx"'
    return response
# -------------------
# VUE SYNTHÈSE
# -------------------

@login_required
@user_passes_test(enseignant_required)
def synthese_enseignant(request):
    """Vue synthèse pour l'enseignant"""
    cours = request.user.cours_enseignant.all()
    classes = Classe.objects.filter(cours__enseignant=request.user).distinct()
    
    # Statistiques globales
    stats = {
        'total_cours': cours.count(),
        'total_classes': classes.count(),
        'total_etudiants': Etudiant.objects.filter(classe__in=classes).count(),
        'total_seances': Seance.objects.filter(cours__enseignant=request.user).count(),
        'presences_aujourdhui': Presence.objects.filter(
            seance__cours__enseignant=request.user,
            seance__date=timezone.now().date(),
            statut__in=['present', 'retard']
        ).count(),
        'cours_par_classe': []
    }
    
    for classe in classes:
        cours_classe = cours.filter(classe=classe)
        stats['cours_par_classe'].append({
            'classe': classe,
            'nombre_cours': cours_classe.count(),
            'nombre_etudiants': classe.etudiants.count(),
            'nombre_seances': Seance.objects.filter(cours__in=cours_classe).count()
        })
    
    # Dernières séances
    dernieres_seances = Seance.objects.filter(
        cours__enseignant=request.user
    ).order_by('-date', '-heure_debut')[:10]
    
    # Prochaines séances (aujourd'hui et après)
    prochaines_seances = Seance.objects.filter(
        cours__enseignant=request.user,
        date__gte=timezone.now().date()
    ).order_by('date', 'heure_debut')[:5]
    
    return render(request, "core/synthese_enseignant.html", {
        "stats": stats,
        "dernieres_seances": dernieres_seances,
        "prochaines_seances": prochaines_seances,
        "cours": cours[:5]
    })

# -------------------
# ADMIN CRUD
# -------------------

@login_required
@user_passes_test(admin_required)
def enseignant_create(request):
    """Création d'enseignant par admin"""
    if request.method == "POST":
        form = EnseignantForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Enseignant ajouté avec succès")
            return redirect("core:admin_dashboard")
    else:
        form = EnseignantForm()
    return render(request, "core/admin_form.html", {
        "form": form, 
        "title": "Ajouter un enseignant",
        "back_url": "core:admin_dashboard"
    })

@login_required
@user_passes_test(admin_required)
def enseignant_update(request, pk):
    """Modification d'enseignant par admin"""
    enseignant = get_object_or_404(User, pk=pk, role="enseignant")
    if request.method == "POST":
        form = EnseignantForm(request.POST, instance=enseignant)
        if form.is_valid():
            form.save()
            messages.success(request, "Enseignant modifié avec succès")
            return redirect("core:admin_dashboard")
    else:
        form = EnseignantForm(instance=enseignant)
    return render(request, "core/admin_form.html", {
        "form": form, 
        "title": "Modifier l'enseignant",
        "back_url": "core:admin_dashboard"
    })

@login_required
@user_passes_test(admin_required)
def enseignant_delete(request, pk):
    """Suppression d'enseignant par admin"""
    enseignant = get_object_or_404(User, pk=pk, role="enseignant")
    if request.method == "POST":
        enseignant.delete()
        messages.success(request, "Enseignant supprimé avec succès")
        return redirect("core:admin_dashboard")
    return render(request, "core/admin_confirm_delete.html", {
        "object": enseignant,
        "object_type": "enseignant",
        "back_url": "core:admin_dashboard"
    })

# ... (les autres fonctions admin restent similaires mais avec des améliorations)

# -------------------
# API & DONNÉES
# -------------------
@login_required
def api_stats_cours(request, cours_id):
    """API pour les statistiques d'un cours (JSON)"""
    cours = get_object_or_404(Cours, id=cours_id, enseignant=request.user)
    
    # Calculer le taux de présence
    total_seances = cours.seances.count()
    if total_seances > 0:
        total_presences = Presence.objects.filter(seance__cours=cours).count()
        presentes = Presence.objects.filter(seance__cours=cours, statut__in=['present', 'retard']).count()
        taux_presence = (presentes / total_presences) * 100 if total_presences > 0 else 0
    else:
        taux_presence = 0
    
    data = {
        'cours': cours.nom,
        'classe': cours.classe.nom,
        'taux_presence': round(taux_presence, 1)
    }
    
    return JsonResponse(data)

@login_required
def api_presences_seance(request, seance_id):
    """API pour les présences d'une séance"""
    seance = get_object_or_404(Seance, id=seance_id, cours__enseignant=request.user)
    presences = Presence.objects.filter(seance=seance).select_related('etudiant')
    
    data = {
        'seance': {
            'id': seance.id,
            'date': seance.date,
            'cours': seance.cours.nom
        },
        'presences': [
            {
                'etudiant': {
                    'id': p.etudiant.id,
                    'nom': p.etudiant.nom,
                    'prenom': p.etudiant.prenom
                },
                'statut': p.statut
            } for p in presences
        ]
    }
    
    return JsonResponse(data)

# -------------------------------
# ADMIN CRUD - LISTES
# -------------------------------

@login_required
@user_passes_test(admin_required)
def enseignant_list(request):
    """Liste des enseignants (admin)"""
    enseignants = User.objects.filter(role="enseignant").order_by('username')
    
    # Recherche
    query = request.GET.get('q')
    if query:
        enseignants = enseignants.filter(
            Q(username__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query)
        )
    
    paginator = Paginator(enseignants, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, "core/admin/enseignant_list.html", {
        "enseignants": page_obj,
        "query": query
    })

@login_required
@user_passes_test(admin_required)
def classe_list(request):
    """Liste des classes (admin)"""
    classes = Classe.objects.all().order_by('nom')
    
    # Recherche
    query = request.GET.get('q')
    if query:
        classes = classes.filter(nom__icontains=query)
    
    # Statistiques
    classes = classes.annotate(
        nb_etudiants=Count('etudiants'),
        nb_cours=Count('cours')
    )
    paginator = Paginator(classes, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    # Ajouter ces statistiques
    total_etudiants = Etudiant.objects.count()
    total_cours = Cours.objects.count()
    
    return render(request, "core/admin/classe_list.html", {
        "classes": page_obj,
        "query": query,
        "total_etudiants": total_etudiants,
        "total_cours": total_cours
    })

@login_required
@user_passes_test(admin_required)
def etudiant_list(request):
    """Liste des étudiants (admin)"""
    etudiants = Etudiant.objects.all().select_related('classe').order_by('classe__nom', 'nom', 'prenom')
    
    # Filtres
    classe_id = request.GET.get('classe')
    query = request.GET.get('q')
    
    if classe_id:
        etudiants = etudiants.filter(classe_id=classe_id)
    
    if query:
        etudiants = etudiants.filter(
            Q(nom__icontains=query) |
            Q(prenom__icontains=query) |
            Q(matricule__icontains=query) |
            Q(classe__nom__icontains=query)
        )
    
    paginator = Paginator(etudiants, 30)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    classes = Classe.objects.all()
    
    return render(request, "core/admin/etudiant_list.html", {
        "etudiants": page_obj,
        "classes": classes,
        "classe_selected": int(classe_id) if classe_id else None,
        "classe_nom": Classe.objects.get(id=classe_id).nom if classe_id else "Toutes les classes",
        "query": query
    })

@login_required
@user_passes_test(admin_required)
def admin_cours_list(request):
    """Liste des cours (admin)"""
    cours = Cours.objects.all().select_related('enseignant', 'classe').order_by('-created_at')
    
    # Filtres
    enseignant_id = request.GET.get('enseignant')
    classe_id = request.GET.get('classe')
    query = request.GET.get('q')
    
    if enseignant_id:
        cours = cours.filter(enseignant_id=enseignant_id)
    
    if classe_id:
        cours = cours.filter(classe_id=classe_id)
    
    if query:
        cours = cours.filter(
            Q(nom__icontains=query) |
            Q(enseignant__username__icontains=query) |
            Q(classe__nom__icontains=query)
        )
    
    paginator = Paginator(cours, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    enseignants = User.objects.filter(role="enseignant")
    classes = Classe.objects.all()
    # Ajouter ces statistiques
    cours_actifs = Cours.objects.filter(seances__isnull=False).distinct().count()
    total_seances = Seance.objects.count()
    enseignants_distincts = User.objects.filter(role="enseignant", cours__isnull=False).distinct().count()
    
    return render(request, "core/admin/cours_list.html", {
        "cours": page_obj,
        "enseignants": enseignants,
        "classes": classes,
        "enseignant_selected": int(enseignant_id) if enseignant_id else None,
        "classe_selected": int(classe_id) if classe_id else None,
        "query": query,
        "cours_actifs": cours_actifs,
        "total_seances": total_seances,
        "enseignants_distincts": enseignants_distincts
    })

# -------------------------------
# ADMIN CRUD - COURS (spécifiques admin)
# -------------------------------

@login_required
@user_passes_test(admin_required)
def admin_cours_create(request):
    """Création de cours par admin"""
    if request.method == "POST":
        form = CoursForm(request.POST)
        if form.is_valid():
            cours = form.save()
            messages.success(request, f"Cours '{cours.nom}' créé avec succès !")
            return redirect("admin_cours_list")
    else:
        form = CoursForm()
    
    return render(request, "core/admin/cours_form.html", {
        "form": form,
        "title": "Créer un cours",
        "action_url": "admin_cours_create"
    })

@login_required
@user_passes_test(admin_required)
def admin_cours_update(request, pk):
    """Modification de cours par admin"""
    cours = get_object_or_404(Cours, pk=pk)
    
    if request.method == "POST":
        form = CoursForm(request.POST, instance=cours)
        if form.is_valid():
            form.save()
            messages.success(request, f"Cours '{cours.nom}' modifié avec succès !")
            return redirect("admin_cours_list")
    else:
        form = CoursForm(instance=cours)
    
    return render(request, "core/admin/cours_form.html", {
        "form": form,
        "title": f"Modifier le cours : {cours.nom}",
        "action_url": "admin_cours_update"
    })

@login_required
@user_passes_test(admin_required)
def admin_cours_delete(request, pk):
    """Suppression de cours par admin"""
    cours = get_object_or_404(Cours, pk=pk)
    
    if request.method == "POST":
        nom_cours = cours.nom
        cours.delete()
        messages.success(request, f"Cours '{nom_cours}' supprimé avec succès !")
        return redirect("admin_cours_list")
    
    return render(request, "core/admin/cours_confirm_delete.html", {"cours": cours})

# -------------------------------
# ADMIN IMPORT ÉTUDIANTS
# -------------------------------

@login_required
@user_passes_test(admin_required)
def admin_importer_etudiants(request):
    """Import d'étudiants par admin"""
    if request.method == "POST":
        form = ImportEtudiantsForm(request.POST, request.FILES)
        if form.is_valid():
            fichier = request.FILES['fichier']
            classe = form.cleaned_data['classe']
            mode = form.cleaned_data.get('mode', 'create')
            
            try:
                if fichier.name.endswith('.xlsx'):
                    wb = openpyxl.load_workbook(fichier)
                    ws = wb.active
                    lignes = list(ws.iter_rows(values_only=True))
                else:
                    data = fichier.read().decode('utf-8')
                    io_string = io.StringIO(data)
                    lignes = list(csv.reader(io_string))
                
                etudiants_importes = 0
                etudiants_modifies = 0
                errors = []
                
                for i, row in enumerate(lignes[1:], start=2):
                    try:
                        if len(row) >= 2 and row[0] and row[1]:
                            matricule = str(row[0]).strip()
                            nom = str(row[1]).strip()
                            prenom = str(row[2]).strip() if len(row) > 2 else ""
                            
                            if mode == 'create':
                                etudiant, created = Etudiant.objects.get_or_create(
                                    matricule=matricule,
                                    defaults={'nom': nom, 'prenom': prenom, 'classe': classe}
                                )
                                if created:
                                    etudiants_importes += 1
                            else:
                                etudiant, created = Etudiant.objects.update_or_create(
                                    matricule=matricule,
                                    defaults={'nom': nom, 'prenom': prenom, 'classe': classe}
                                )
                                if created:
                                    etudiants_importes += 1
                                else:
                                    etudiants_modifies += 1
                    except Exception as e:
                        errors.append(f"Ligne {i}: {str(e)}")
                
                if errors:
                    messages.warning(request, f"Import terminé avec {len(errors)} erreurs")
                else:
                    msg = f"{etudiants_importes} étudiants importés"
                    if mode == 'update':
                        msg += f", {etudiants_modifies} modifiés"
                    messages.success(request, msg)
                
                return redirect("etudiant_list")
                
            except Exception as e:
                messages.error(request, f"Erreur lors de la lecture du fichier: {str(e)}")
    else:
        form = ImportEtudiantsForm()
    
    return render(request, "core/admin/importer_etudiants.html", {"form": form})

# -------------------------------
# FONCTIONNALITÉS DE RECHERCHE
# -------------------------------

@login_required
def recherche_etudiants(request):
    """Recherche d'étudiants"""
    query = request.GET.get('q', '')
    results = []
    
    if query:
        results = Etudiant.objects.filter(
            Q(nom__icontains=query) |
            Q(prenom__icontains=query) |
            Q(matricule__icontains=query) |
            Q(classe__nom__icontains=query)
        ).select_related('classe')[:10]
    
    return render(request, "core/recherche_etudiants.html", {
        "results": results,
        "query": query
    })

@login_required
def recherche_cours(request):
    """Recherche de cours"""
    query = request.GET.get('q', '')
    results = []
    
    if request.user.role == "admin":
        queryset = Cours.objects.all()
    else:
        queryset = request.user.cours_enseignant.all()
    
    if query:
        results = queryset.filter(
            Q(nom__icontains=query) |
            Q(classe__nom__icontains=query) |
            Q(enseignant__username__icontains=query)
        ).select_related('enseignant', 'classe')[:10]
    
    return render(request, "core/recherche_cours.html", {
        "results": results,
        "query": query
    })

@login_required
def recherche_seances(request):
    """Recherche de séances"""
    query = request.GET.get('q', '')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    results = []
    
    queryset = Seance.objects.filter(cours__enseignant=request.user)
    
    if query:
        queryset = queryset.filter(
            Q(cours__nom__icontains=query) |
            Q(description__icontains=query)
        )
    
    if date_debut:
        queryset = queryset.filter(date__gte=date_debut)
    
    if date_fin:
        queryset = queryset.filter(date__lte=date_fin)
    
    results = queryset.select_related('cours', 'cours__classe').order_by('-date')[:20]
    
    return render(request, "core/recherche_seances.html", {
        "results": results,
        "query": query,
        "date_debut": date_debut,
        "date_fin": date_fin
    })

# -------------------------------
# VUES API POUR AUTOCOMPLÉTION
# -------------------------------

@login_required
def api_recherche_etudiants(request):
    """API pour autocomplétion étudiants"""
    query = request.GET.get('q', '')
    classe_id = request.GET.get('classe')
    
    queryset = Etudiant.objects.all()
    
    if request.user.role == "enseignant":
        classes_enseignant = Classe.objects.filter(cours__enseignant=request.user).distinct()
        queryset = queryset.filter(classe__in=classes_enseignant)
    
    if classe_id:
        queryset = queryset.filter(classe_id=classe_id)
    
    if query:
        queryset = queryset.filter(
            Q(nom__icontains=query) |
            Q(prenom__icontains=query) |
            Q(matricule__icontains=query)
        )[:10]
    
    results = [
        {
            'id': etu.id,
            'text': f"{etu.nom} {etu.prenom or ''} ({etu.matricule}) - {etu.classe.nom}",
            'matricule': etu.matricule,
            'nom': etu.nom,
            'prenom': etu.prenom,
            'classe': etu.classe.nom
        }
        for etu in queryset
    ]
    
    return JsonResponse({'results': results})

@login_required
def api_recherche_cours(request):
    """API pour autocomplétion cours"""
    query = request.GET.get('q', '')
    
    if request.user.role == "admin":
        queryset = Cours.objects.all()
    else:
        queryset = request.user.cours_enseignant.all()
    
    if query:
        queryset = queryset.filter(
            Q(nom__icontains=query) |
            Q(classe__nom__icontains=query)
        )[:10]
    
    results = [
        {
            'id': cours.id,
            'text': f"{cours.nom} - {cours.classe.nom}",
            'nom': cours.nom,
            'classe': cours.classe.nom,
            'enseignant': cours.enseignant.username
        }
        for cours in queryset.select_related('classe', 'enseignant')
    ]
    
    return JsonResponse({'results': results})


# -------------------
# ADMIN CRUD - CLASSES
# -------------------

@login_required
@user_passes_test(admin_required)
def classe_create(request):
    """Création de classe par admin"""
    if request.method == "POST":
        form = ClasseForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Classe créée avec succès !")
            return redirect("core:admin_dashboard")
    else:
        form = ClasseForm()
    
    return render(request, "core/admin_form.html", {
        "form": form,
        "title": "Créer une classe",
        "back_url": "core:admin_dashboard"
    })

@login_required
@user_passes_test(admin_required)
def classe_update(request, pk):
    """Modification de classe par admin"""
    classe = get_object_or_404(Classe, pk=pk)
    
    if request.method == "POST":
        form = ClasseForm(request.POST, instance=classe)
        if form.is_valid():
            form.save()
            messages.success(request, f"Classe '{classe.nom}' modifiée avec succès !")
            return redirect("core:admin_dashboard")
    else:
        form = ClasseForm(instance=classe)
    
    return render(request, "core/admin_form.html", {
        "form": form,
        "title": f"Modifier la classe : {classe.nom}",
        "back_url": "core:admin_dashboard"
    })

@login_required
@user_passes_test(admin_required)
def classe_delete(request, pk):
    """Suppression de classe par admin"""
    classe = get_object_or_404(Classe, pk=pk)
    
    if request.method == "POST":
        nom_classe = classe.nom
        classe.delete()
        messages.success(request, f"Classe '{nom_classe}' supprimée avec succès !")
        return redirect("core:admin_dashboard")
    
    return render(request, "core/admin_confirm_delete.html", {
        "object": classe,
        "object_type": "classe",
        "back_url": "core:admin_dashboard"
    })

# -------------------
# ADMIN CRUD - ÉTUDIANTS
# -------------------

@login_required
@user_passes_test(admin_required)
def etudiant_create(request):
    """Création d'étudiant par admin"""
    if request.method == "POST":
        form = EtudiantForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Étudiant créé avec succès !")
            return redirect("core:admin_dashboard")
    else:
        form = EtudiantForm()
    
    return render(request, "core/admin_form.html", {
        "form": form,
        "title": "Créer un étudiant",
        "back_url": "core:admin_dashboard"
    })

@login_required
@user_passes_test(admin_required)
def etudiant_update(request, pk):
    """Modification d'étudiant par admin"""
    etudiant = get_object_or_404(Etudiant, pk=pk)
    
    if request.method == "POST":
        form = EtudiantForm(request.POST, instance=etudiant)
        if form.is_valid():
            form.save()
            messages.success(request, f"Étudiant '{etudiant.nom}' modifié avec succès !")
            return redirect("core:admin_dashboard")
    else:
        form = EtudiantForm(instance=etudiant)
    
    return render(request, "core/admin_form.html", {
        "form": form,
        "title": f"Modifier l'étudiant : {etudiant.nom}",
        "back_url": "core:admin_dashboard"
    })

@login_required
@user_passes_test(admin_required)
def etudiant_delete(request, pk):
    """Suppression d'étudiant par admin"""
    etudiant = get_object_or_404(Etudiant, pk=pk)
    
    if request.method == "POST":
        nom_etudiant = etudiant.nom
        etudiant.delete()
        messages.success(request, f"Étudiant '{nom_etudiant}' supprimé avec succès !")
        return redirect("core:admin_dashboard")
    
    return render(request, "core/admin_confirm_delete.html", {
        "object": etudiant,
        "object_type": "étudiant",
        "back_url": "core:admin_dashboard"
    })


